import requests
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
import re
import time
import openpyxl
import configparser
import os


recommendAcountsFile = open('RecETH.txt', 'r')
recommendAcounts = recommendAcountsFile.readlines()

#myAcountsFile = open('myeth.txt', 'r')
#myAcounts = myAcountsFile.readlines()
config = configparser.ConfigParser()

# 创建配置文件
if not os.path.exists('./config.ini'):
    file = open('config.ini','w')
    file.close()
config.read('config.ini')
file = open('config.ini','w+')
if not config.has_section('RunConfig'):
    config.read('config.ini')
    file = open('config.ini','w+')
    config.add_section('RunConfig')
    config.set('RunConfig', 'CurrentLine', '74000')
    config.write(file)
    file.close()
linesCount = int(config.get('RunConfig', 'CurrentLine'))

def GetDriver(url, ethAddress):
    for trytime in range(1, 4):
        try:
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--disable-gpu')

            # 创建浏览器对象
            driver = webdriver.Chrome(executable_path=r'E:\VS2015\NFTRecommend\env\Scripts\chromedriver.exe', chrome_options=chrome_options)
            driver.get(url)
            driver.implicitly_wait(10)
            elem = driver.find_element_by_class_name('input.form-control')
            elem.send_keys(ethAddress)
            elem.submit()
            time.sleep(2)
            driver.refresh()
            return driver
        except:
            continue
    return None

    #try:
       # element_submmit = driver.find_element_by_class_name('btn.btn-fill.py-2').submit()
   # except:
        #print("已生成链接。")
workbook = openpyxl.load_workbook('NFT.xlsx')
worksheet = workbook.get_sheet_by_name('Sheet1')
maxRow = worksheet.max_row
while True:
    for rowIndex in range(1, maxRow + 1):
        acount = worksheet.cell(row=rowIndex,column=1).value
        withdrawOrNot = worksheet.cell(row=rowIndex,column=3).value

        # 已经提取，直接下一个
        if withdrawOrNot == '1':
            continue
        
        config.read('config.ini')
        file = open('config.ini','w+')

        driverMy = GetDriver('https://nftsea.one/', acount)
        if driverMy is None:
            print("我的地址【%s】创建浏览器对象失败！" % acount.strip())
            continue

        # print(driverMy.page_source)
        searchObj = re.search(r'text-primary copy" text="(https:.*?)"', driverMy.page_source, re.I)
        if searchObj is None:
            print("我的地址【%s】分享链接失败, 未找到分享链接！" % acount.strip())
            driverMy.quit()
            continue
        myRecommendLink = searchObj.group(1)

        # 获取推荐数
        RecommCount = 0
        searchObj = re.search(r'heading-h5 text-primary">(\d{1,2})', driverMy.page_source, re.I)
        if searchObj is not None:
            RecommCount = searchObj.group(1)
        remainRecomm = 50 - int(RecommCount)

        pattern = re.compile(r'class="text\-primary">(\d+)<\/span> NS', re.I)
        patternlist = pattern.findall(driverMy.page_source)

        # 获取余额
        remainCoin = int(patternlist[0])

        # 获取奖励数
        rewardCoin = int(patternlist[1])

        if remainCoin < 50 and rewardCoin > 460:
            worksheet.cell(row=rowIndex, column=2, value=rewardCoin)
            worksheet.cell(row=rowIndex, column=3, value='1')
            print("地址【%s】已提币！" % acount.strip())
            workbook.save('NFT.xlsx')
            continue

        # 已推荐满，但是还未提取，赶紧提取
        if remainRecomm == 0:
            if remainCoin >= 50:
                print("我的地址【%s】推荐已满！可以提取， 正在提取中..." % acount.strip())
                js = 'document.getElementsByClassName("btn btn-fill")[0].click()'
                driverMy.execute_script(js)
                print("我的地址【%s】提取完成。" % acount.strip())
                worksheet.cell(row=rowIndex, column=2, value=rewardCoin)
                worksheet.cell(row=rowIndex, column=3, value='1')
                driverMy.quit()
                workbook.save('NFT.xlsx')
            continue

        driverMy.quit()
        print("正在分享我的第%s个地址【%s】, 分享链接为【%s】" % (rowIndex, acount.strip(), myRecommendLink))

        successCount = 0
        while successCount < remainRecomm:
            recommendETH = recommendAcounts[linesCount]
            driverRecomm = GetDriver(myRecommendLink, recommendETH)
            if driverRecomm is None:
                print("     【%s】号地址推荐失败：创建浏览器对象失败，其地址为【%s】" % (linesCount, recommendETH.strip()))
                continue
            searchObj = re.search(r'text-primary copy" text="(https:.*?)"', driverRecomm.page_source, re.I)
            if searchObj is None:
                print("     【%s】号地址推荐失败，其地址为【%s】" % (linesCount, recommendETH.strip()))
                driverRecomm.quit()
                continue
            print("     正在推荐第%s个地址，推荐成功！其推荐链接为【%s】" % (linesCount, searchObj.group(1)))
            successCount = successCount + 1
            linesCount = linesCount - 1
            config.set('RunConfig', 'CurrentLine', str(linesCount))
            driverRecomm.quit()
        config.write(file)
        file.close()
        